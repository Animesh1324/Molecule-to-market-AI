"""Ingest and query secondary market data (IQVIA/IMS TSA, PharmaTrac, AWACS).

Why this exists: public sources answer "what molecules exist". They do not
answer "who am I competing against in this market next quarter" — that lives in
a syndicated audit extract a brand team licenses. This module turns such an
extract into the competitor set, company leaderboard, and molecule market size
the planning modules read.

Two properties matter more than speed here:

* **Provenance.** Every returned number carries the dataset it came from, so a
  reviewer can trace a share figure back to a file and a period. Nothing is
  synthesised; a molecule absent from the extract returns empty, not a guess.
* **Re-ingestion safety.** Uploading a newer period replaces the prior dataset
  for the same file wholesale, inside one transaction. A half-written refresh
  would silently mix two periods into one market size.

Column names differ across vendors and even across months of the same vendor,
so headers are matched against alias lists rather than fixed positions.
"""
from __future__ import annotations

import csv
import logging
import os
import re
import uuid
from datetime import datetime
from typing import Any, Dict, Iterable, Iterator, List, Optional, Tuple

from sqlalchemy import and_, func, or_

from ..db.database import SessionLocal
from ..db.manual_competitor_models import ManualCompetitorORM
from ..db.market_models import MarketBrandORM, MarketDatasetORM

logger = logging.getLogger(__name__)

# --------------------------------------------------------------------------
# Header mapping
# --------------------------------------------------------------------------

# Ordered by specificity: the first alias found in the header row wins, so an
# exact vendor column beats a looser synonym.
COLUMN_ALIASES: Dict[str, List[str]] = {
    "molecule": ["MOLECULE_DESC", "MOLECULE DESC", "MOLECULE", "MOLECULE NAME",
                 "COMPOSITION", "GENERIC NAME", "GENERIC", "SALT"],
    "brand": ["BRANDS", "BRAND", "BRAND NAME", "PRODUCT", "PRODUCT NAME"],
    "company": ["COMPANY", "CORPORATE", "CORPORATION", "COMPANY NAME"],
    "manufacturer": ["MANUFACT. DESC", "MANUFACTURER", "MANUFACTURER NAME", "MFR", "MFG"],
    "ownership": ["INDIAN_MNC", "INDIAN/MNC", "ORIGIN", "COMPANY TYPE"],
    "subgroup": ["SUBGROUP", "SUB GROUP", "SUB-GROUP"],
    "group": ["GROUP", "THERAPY GROUP", "THERAPEUTIC GROUP"],
    "supergroup": ["SUPERGROUP", "SUPER GROUP", "THERAPY AREA", "THERAPEUTIC AREA"],
    "acute_chronic": ["ACUTE_CHRONIC", "ACUTE/CHRONIC", "ACUTE CHRONIC"],
    "form": ["SHORT DESCRIPTION", "DOSAGE FORM", "FORM", "NFC 1"],
    "pack": ["PACK_DESC", "PACK DESC", "PACK", "PACK DESCRIPTION", "SKU"],
    "plain_combination": ["PLAIN/COMBINATION", "PLAIN_COMBINATION", "TYPE"],
}

# Salt and hydrate forms that change the source wording but not the competitor.
# Stripped only when they are not the whole name, so "SODIUM CHLORIDE" survives.
_SALT_TOKENS = {
    "SODIUM", "POTASSIUM", "CALCIUM", "MAGNESIUM", "ZINC", "DISODIUM",
    "HYDROCHLORIDE", "HCL", "DIHYDROCHLORIDE", "CHLORIDE", "BROMIDE",
    "SULFATE", "SULPHATE", "BISULFATE", "PHOSPHATE", "DIPHOSPHATE",
    "MALEATE", "TARTRATE", "BITARTRATE", "FUMARATE", "SUCCINATE", "CITRATE",
    "ACETATE", "OXALATE", "NITRATE", "CARBONATE", "GLUCONATE", "LACTATE",
    "BESYLATE", "BESILATE", "MESYLATE", "MESILATE", "TOSYLATE", "TOSILATE",
    "PAMOATE", "STEARATE", "PALMITATE", "PROPIONATE", "VALERATE", "FUROATE",
    "DIPROPIONATE", "HEMIFUMARATE", "HEMIHYDRATE", "MONOHYDRATE", "DIHYDRATE",
    "TRIHYDRATE", "HYDRATE", "ANHYDROUS", "SALT", "SALTS", "BASE",
    "ETHANOLATE", "SOLVATE", "ARGININE", "MEGLUMINE", "TROMETAMOL",
}

_SPLIT_RE = re.compile(r"\s*[+/]\s*|\s+AND\s+|\s*&\s*")
_NON_ALNUM = re.compile(r"[^A-Z0-9]+")
# "MAT AUG'24" / "MAT SEP 2024" — value columns, excluding UNIT/QTY variants.
_MAT_RE = re.compile(r"^MAT\s+([A-Z]{3})[\s']*(\d{2,4})$")
_MONTHS = {m: i for i, m in enumerate(
    ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
     "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"], start=1)}


def normalise_molecule(raw: str) -> str:
    """Salt-stripped, punctuation-free lookup key for a molecule description.

    "ROSUVASTATIN CALCIUM"            -> "ROSUVASTATIN"
    "EMPAGLIFLOZIN + LINAGLIPTIN"     -> "EMPAGLIFLOZIN+LINAGLIPTIN"
    "AMOXICILLIN TRIHYDRATE + CLA..." -> "AMOXICILLIN+CLA..."

    Lossy by design — `molecule_desc` keeps the source wording for display.
    """
    text = (raw or "").upper().strip()
    if not text:
        return ""
    parts: List[str] = []
    for component in _SPLIT_RE.split(text):
        tokens = [t for t in _NON_ALNUM.sub(" ", component).split() if t]
        if not tokens:
            continue
        kept = [t for t in tokens if t not in _SALT_TOKENS]
        # Every token was a salt word ("SODIUM CHLORIDE"): keep the original.
        parts.append("".join(kept) if kept else "".join(tokens))
    return "+".join(p for p in parts if p)


def molecule_search_key(raw: str) -> str:
    """Key for the *query* side — a single molecule, never a combination."""
    normalised = normalise_molecule(raw)
    return normalised.split("+")[0] if normalised else ""


def _period_sort_key(label: str) -> Tuple[int, int]:
    match = _MAT_RE.match(label.strip().upper())
    if not match:
        return (0, 0)
    month, year = match.group(1), match.group(2)
    year_num = int(year)
    if year_num < 100:
        year_num += 2000
    return (year_num, _MONTHS.get(month, 0))


def _resolve_columns(header: List[Any]) -> Tuple[Dict[str, int], List[Tuple[str, int]]]:
    """Map logical field -> column index, plus ordered MAT value columns."""
    seen: Dict[str, int] = {}
    for index, cell in enumerate(header):
        if cell is None:
            continue
        key = str(cell).strip().upper()
        if key and key not in seen:
            seen[key] = index

    mapping: Dict[str, int] = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in seen:
                mapping[field] = seen[alias]
                break

    # Value columns: "MAT <MON><YY>" only. UNIT/QTY MAT columns are volume, not
    # value, and mixing them into the same field would corrupt market share.
    periods: List[Tuple[str, int]] = []
    for label, index in seen.items():
        if label.startswith(("UNIT", "QTY")):
            continue
        if _MAT_RE.match(label):
            periods.append((label, index))
    periods.sort(key=lambda pair: _period_sort_key(pair[0]), reverse=True)

    if not periods:
        # Some extracts label the latest period without the MAT prefix; fall
        # back to any numeric-looking trailing column the aliases did not claim.
        for label, index in seen.items():
            if label.startswith(("VALUE", "SALES", "MAT")):
                periods.append((label, index))
        periods.sort(key=lambda pair: _period_sort_key(pair[0]), reverse=True)

    unit_index = seen.get("UNIT MAT " + periods[0][0][4:]) if periods else None
    if unit_index is not None:
        mapping["units_latest"] = unit_index
    return mapping, periods


def _to_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace(",", "").strip()
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _clean(value: Any, limit: int = 240) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text[:limit] or None


# --------------------------------------------------------------------------
# Row readers
# --------------------------------------------------------------------------

def _iter_xlsx(path: str) -> Iterator[List[Any]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        # Prefer a sheet that looks like a flat extract over a pivot cache.
        sheet = None
        for name in workbook.sheetnames:
            if name.strip().upper() in {"DATA", "RAW", "BASE", "SHEET1"}:
                sheet = workbook[name]
                break
        if sheet is None:
            sheet = max(workbook.worksheets, key=lambda ws: (ws.max_row or 0))
        for row in sheet.iter_rows(values_only=True):
            yield list(row)
    finally:
        workbook.close()


def _iter_delimited(path: str) -> Iterator[List[Any]]:
    delimiter = "\t" if path.lower().endswith((".tsv", ".tab")) else ","
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as handle:
        for row in csv.reader(handle, delimiter=delimiter):
            yield row


def _row_iterator(path: str) -> Iterator[List[Any]]:
    extension = os.path.splitext(path)[1].lower()
    if extension in {".xlsx", ".xlsm", ".xls"}:
        return _iter_xlsx(path)
    if extension in {".csv", ".tsv", ".tab", ".txt"}:
        return _iter_delimited(path)
    raise ValueError(f"{extension or 'This file type'} cannot be parsed as a market extract.")


def looks_like_market_extract(path: str) -> bool:
    """Cheap pre-check so an unrelated upload is not ingested as market data."""
    try:
        rows = _row_iterator(path)
        for row in rows:
            mapping, periods = _resolve_columns(list(row))
            required = {"molecule", "brand"}
            return required.issubset(mapping) and bool(periods)
    except Exception:
        return False
    return False


# --------------------------------------------------------------------------
# Ingestion
# --------------------------------------------------------------------------

BATCH = 2000


def ingest_market_file(
    path: str,
    original_filename: Optional[str] = None,
    source_label: str = "Secondary data",
    market: str = "India",
    value_unit: str = "INR Cr",
    project_id: Optional[str] = None,
    upload_file_id: Optional[str] = None,
    replace_existing: bool = True,
) -> Dict[str, Any]:
    """Parse one extract into `market_brands` and register it as a dataset.

    Re-ingesting a file with the same original filename replaces the previous
    dataset, so refreshing a period never leaves two periods double-counted in
    the same market size.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    filename = original_filename or os.path.basename(path)
    rows = _row_iterator(path)

    header: Optional[List[Any]] = None
    mapping: Dict[str, int] = {}
    periods: List[Tuple[str, int]] = []

    # The header is not always row 1 — vendor exports carry title banners.
    for candidate in rows:
        mapping, periods = _resolve_columns(list(candidate))
        if {"molecule", "brand"}.issubset(mapping) and periods:
            header = list(candidate)
            break
    if header is None:
        raise ValueError(
            "No molecule/brand/value columns found. Expected columns such as "
            "MOLECULE_DESC, BRANDS, COMPANY and a MAT value column."
        )

    period_labels = [label for label, _ in periods[:3]]
    dataset_id = uuid.uuid4().hex
    session = SessionLocal()
    written = 0
    brands: set = set()
    molecules: set = set()
    companies: set = set()
    total_value = 0.0
    batch: List[MarketBrandORM] = []

    def flush() -> None:
        if batch:
            session.bulk_save_objects(batch)
            batch.clear()

    try:
        superseded: List[str] = []
        if replace_existing:
            superseded = [
                row.id for row in session.query(MarketDatasetORM.id)
                .filter(MarketDatasetORM.original_filename == filename).all()
            ]

        molecule_col = mapping["molecule"]
        brand_col = mapping["brand"]

        for raw in rows:
            row = list(raw)
            if len(row) <= max(molecule_col, brand_col):
                continue
            molecule_desc = _clean(row[molecule_col])
            brand = _clean(row[brand_col], 160)
            if not molecule_desc or not brand:
                continue

            def cell(field: str, limit: int = 240) -> Optional[str]:
                index = mapping.get(field)
                if index is None or index >= len(row):
                    return None
                return _clean(row[index], limit)

            values: List[Optional[float]] = []
            for _, index in periods[:3]:
                values.append(_to_float(row[index]) if index < len(row) else None)
            while len(values) < 3:
                values.append(None)

            key = normalise_molecule(molecule_desc)
            latest = values[0] or 0.0
            units_index = mapping.get("units_latest")

            batch.append(MarketBrandORM(
                id=uuid.uuid4().hex,
                dataset_id=dataset_id,
                molecule_desc=molecule_desc,
                molecule_key=key,
                is_combination=1 if "+" in key else 0,
                brand=brand,
                company=cell("company", 160),
                manufacturer=cell("manufacturer", 160),
                ownership=cell("ownership", 40),
                subgroup=cell("subgroup"),
                group_name=cell("group"),
                supergroup=cell("supergroup"),
                acute_chronic=cell("acute_chronic", 40),
                dosage_form=cell("form"),
                pack_desc=cell("pack"),
                value_latest=latest,
                value_prev=values[1],
                value_prev2=values[2],
                units_latest=(_to_float(row[units_index])
                              if units_index is not None and units_index < len(row) else None),
                period_latest=period_labels[0] if period_labels else None,
                period_prev=period_labels[1] if len(period_labels) > 1 else None,
                period_prev2=period_labels[2] if len(period_labels) > 2 else None,
            ))
            written += 1
            brands.add(brand.upper())
            molecules.add(key)
            if batch and len(batch) >= BATCH:
                flush()
            company = cell("company", 160)
            if company:
                companies.add(company.upper())
            total_value += latest

        flush()

        if superseded:
            session.query(MarketBrandORM).filter(
                MarketBrandORM.dataset_id.in_(superseded)).delete(synchronize_session=False)
            session.query(MarketDatasetORM).filter(
                MarketDatasetORM.id.in_(superseded)).delete(synchronize_session=False)

        session.add(MarketDatasetORM(
            id=dataset_id,
            original_filename=filename,
            source_label=source_label,
            market=market,
            period_label=period_labels[0] if period_labels else None,
            value_unit=value_unit,
            row_count=written,
            brand_count=len(brands),
            molecule_count=len(molecules),
            company_count=len(companies),
            total_value=round(total_value, 3),
            project_id=project_id,
            upload_file_id=upload_file_id,
            status="ready",
            message=None,
            ingested_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ))
        session.commit()
    except Exception:
        session.rollback()
        logger.exception("Market ingestion failed for %s", filename)
        raise
    finally:
        session.close()

    logger.info("Ingested %s: %d rows, %d brands, %d molecules",
                filename, written, len(brands), len(molecules))
    return {
        "dataset_id": dataset_id,
        "original_filename": filename,
        "period_label": period_labels[0] if period_labels else None,
        "periods": period_labels,
        "rows_ingested": written,
        "brands": len(brands),
        "molecules": len(molecules),
        "companies": len(companies),
        "total_value": round(total_value, 2),
        "value_unit": value_unit,
        "replaced_datasets": len(superseded) if replace_existing else 0,
    }


# --------------------------------------------------------------------------
# Queries
# --------------------------------------------------------------------------

def _growth(latest: Optional[float], previous: Optional[float]) -> Optional[float]:
    """Period-on-period growth %, or None when the base is absent or zero."""
    if latest is None or not previous:
        return None
    return round(((latest - previous) / previous) * 100.0, 1)


def has_market_data() -> bool:
    session = SessionLocal()
    try:
        return session.query(MarketDatasetORM).count() > 0
    finally:
        session.close()


def list_datasets() -> List[Dict[str, Any]]:
    session = SessionLocal()
    try:
        rows = (session.query(MarketDatasetORM)
                .order_by(MarketDatasetORM.ingested_at.desc()).all())
        return [{
            "id": r.id,
            "original_filename": r.original_filename,
            "source_label": r.source_label,
            "market": r.market,
            "period_label": r.period_label,
            "value_unit": r.value_unit,
            "row_count": r.row_count,
            "brand_count": r.brand_count,
            "molecule_count": r.molecule_count,
            "company_count": r.company_count,
            "total_value": r.total_value,
            "project_id": r.project_id,
            "upload_file_id": r.upload_file_id,
            "status": r.status,
            "message": r.message,
            "ingested_at": r.ingested_at,
        } for r in rows]
    finally:
        session.close()


def delete_dataset(dataset_id: str) -> int:
    session = SessionLocal()
    try:
        removed = (session.query(MarketBrandORM)
                   .filter(MarketBrandORM.dataset_id == dataset_id)
                   .delete(synchronize_session=False))
        session.query(MarketDatasetORM).filter(
            MarketDatasetORM.id == dataset_id).delete(synchronize_session=False)
        session.commit()
        return removed
    finally:
        session.close()


def _molecule_filter(molecule: str):
    """Match a molecule inside plain rows and combination rows alike.

    LIKE '%KEY%' rather than equality, because "EMPAGLIFLOZIN+LINAGLIPTIN" is a
    real competitor for Empagliflozin and an equality match would drop it.
    """
    key = molecule_search_key(molecule)
    if not key:
        return None, ""
    return MarketBrandORM.molecule_key.like(f"%{key}%"), key


def _dataset_recency(dataset: MarketDatasetORM) -> Tuple[int, int, str]:
    """Order datasets newest-first by the period they report, not by upload time.

    A team can load an older extract after a newer one; the period the file
    covers is what decides which numbers are current.
    """
    year, month = _period_sort_key(dataset.period_label or "")
    return (year, month, dataset.ingested_at or "")


def active_dataset_id(session, condition) -> Optional[str]:
    """The one dataset a molecule's figures should be read from.

    Aggregating across datasets would sum different periods into a single
    "market size" — MAT JUN'26 plus MAT AUG'24 is not a market, it is a
    double-count, and the same brand would appear twice. So exactly one dataset
    answers any given molecule: the most recent period that actually carries
    rows for it. Older extracts stay queryable for other molecules they cover.
    """
    dataset_ids = [
        row.dataset_id for row in
        session.query(MarketBrandORM.dataset_id).filter(condition).distinct().all()
    ]
    if not dataset_ids:
        return None
    datasets = (session.query(MarketDatasetORM)
                .filter(MarketDatasetORM.id.in_(dataset_ids)).all())
    if not datasets:
        return None
    return max(datasets, key=_dataset_recency).id


def brand_competitors(molecule: str, limit: int = 40) -> Dict[str, Any]:
    """Every brand of a molecule in the market, aggregated from pack rows."""
    condition, key = _molecule_filter(molecule)
    if condition is None:
        return {"molecule": molecule, "brands": [], "market_size": 0.0}

    session = SessionLocal()
    try:
        dataset_id = active_dataset_id(session, condition)
        if dataset_id is None:
            return {"molecule": molecule.title(), "brands": [], "market_size": 0.0,
                    "total_brands": 0, "market_size_prev": 0.0,
                    "market_growth_percent": None, "period": None,
                    "molecule_key": key}
        condition = and_(condition, MarketBrandORM.dataset_id == dataset_id)

        rows = (session.query(
                    MarketBrandORM.brand,
                    MarketBrandORM.company,
                    MarketBrandORM.ownership,
                    MarketBrandORM.molecule_desc,
                    MarketBrandORM.is_combination,
                    MarketBrandORM.subgroup,
                    MarketBrandORM.group_name,
                    MarketBrandORM.acute_chronic,
                    MarketBrandORM.period_latest,
                    func.sum(MarketBrandORM.value_latest).label("value_latest"),
                    func.sum(MarketBrandORM.value_prev).label("value_prev"),
                    func.sum(MarketBrandORM.value_prev2).label("value_prev2"),
                    func.sum(MarketBrandORM.units_latest).label("units_latest"),
                    func.count(MarketBrandORM.id).label("pack_count"),
                )
                .filter(condition)
                .group_by(MarketBrandORM.brand, MarketBrandORM.company)
                .order_by(func.sum(MarketBrandORM.value_latest).desc())
                .all())

        market_size = sum((r.value_latest or 0.0) for r in rows)
        prev_size = sum((r.value_prev or 0.0) for r in rows)

        brands: List[Dict[str, Any]] = []
        for row in rows[:limit]:
            value = row.value_latest or 0.0
            brands.append({
                "brand": row.brand,
                "company": row.company,
                "ownership": row.ownership,
                "molecule_desc": row.molecule_desc,
                "is_combination": bool(row.is_combination),
                "subgroup": row.subgroup,
                "group_name": row.group_name,
                "acute_chronic": row.acute_chronic,
                "value_latest": round(value, 2),
                "value_prev": round(row.value_prev, 2) if row.value_prev else None,
                "growth_percent": _growth(row.value_latest, row.value_prev),
                "units_latest": round(row.units_latest, 1) if row.units_latest else None,
                "market_share_percent": round((value / market_size) * 100.0, 2) if market_size else 0.0,
                "pack_count": row.pack_count,
                "period": row.period_latest,
            })

        return {
            "molecule": molecule.title(),
            "molecule_key": key,
            "brands": brands,
            "total_brands": len(rows),
            "market_size": round(market_size, 2),
            "market_size_prev": round(prev_size, 2),
            "market_growth_percent": _growth(market_size, prev_size),
            "period": rows[0].period_latest if rows else None,
        }
    finally:
        session.close()


def company_leaderboard(molecule: str, limit: int = 15) -> List[Dict[str, Any]]:
    """Corporate share of a molecule's market."""
    condition, _ = _molecule_filter(molecule)
    if condition is None:
        return []
    session = SessionLocal()
    try:
        dataset_id = active_dataset_id(session, condition)
        if dataset_id is None:
            return []
        condition = and_(condition, MarketBrandORM.dataset_id == dataset_id)

        rows = (session.query(
                    MarketBrandORM.company,
                    MarketBrandORM.ownership,
                    func.sum(MarketBrandORM.value_latest).label("value_latest"),
                    func.sum(MarketBrandORM.value_prev).label("value_prev"),
                    func.count(func.distinct(MarketBrandORM.brand)).label("brand_count"),
                )
                .filter(condition, MarketBrandORM.company.isnot(None))
                .group_by(MarketBrandORM.company)
                .order_by(func.sum(MarketBrandORM.value_latest).desc())
                .all())
        total = sum((r.value_latest or 0.0) for r in rows)
        return [{
            "company": r.company,
            "ownership": r.ownership,
            "value_latest": round(r.value_latest or 0.0, 2),
            "growth_percent": _growth(r.value_latest, r.value_prev),
            "brand_count": r.brand_count,
            "market_share_percent": round(((r.value_latest or 0.0) / total) * 100.0, 2) if total else 0.0,
        } for r in rows[:limit]]
    finally:
        session.close()


def class_competitors(molecule: str, limit: int = 12) -> Dict[str, Any]:
    """Rival molecules sold into the same therapeutic group.

    The molecule's own rows set the group; every other molecule in that group is
    a class competitor a brand team argues against, even though it never shares
    an ingredient.
    """
    condition, key = _molecule_filter(molecule)
    if condition is None:
        return {"group": None, "molecules": []}

    session = SessionLocal()
    try:
        dataset_id = active_dataset_id(session, condition)
        if dataset_id is None:
            return {"group": None, "molecules": []}
        condition = and_(condition, MarketBrandORM.dataset_id == dataset_id)

        group_row = (session.query(
                        MarketBrandORM.group_name,
                        func.sum(MarketBrandORM.value_latest).label("value"))
                     .filter(condition, MarketBrandORM.group_name.isnot(None))
                     .group_by(MarketBrandORM.group_name)
                     .order_by(func.sum(MarketBrandORM.value_latest).desc())
                     .first())
        if not group_row:
            return {"group": None, "molecules": []}
        group = group_row.group_name

        rows = (session.query(
                    MarketBrandORM.molecule_key,
                    MarketBrandORM.molecule_desc,
                    func.sum(MarketBrandORM.value_latest).label("value_latest"),
                    func.sum(MarketBrandORM.value_prev).label("value_prev"),
                    func.count(func.distinct(MarketBrandORM.brand)).label("brand_count"),
                )
                .filter(MarketBrandORM.group_name == group,
                        MarketBrandORM.dataset_id == dataset_id)
                .group_by(MarketBrandORM.molecule_key)
                .order_by(func.sum(MarketBrandORM.value_latest).desc())
                .all())

        total = sum((r.value_latest or 0.0) for r in rows)
        molecules: List[Dict[str, Any]] = []
        for row in rows:
            if key and key in (row.molecule_key or ""):
                continue        # the subject molecule is not its own rival
            molecules.append({
                "molecule_key": row.molecule_key,
                "molecule_desc": row.molecule_desc,
                "value_latest": round(row.value_latest or 0.0, 2),
                "growth_percent": _growth(row.value_latest, row.value_prev),
                "brand_count": row.brand_count,
                "class_share_percent": round(((row.value_latest or 0.0) / total) * 100.0, 2) if total else 0.0,
            })
            if len(molecules) >= limit:
                break

        return {
            "group": group,
            "group_value": round(total, 2),
            "molecules": molecules,
        }
    finally:
        session.close()


def top_brands_for_group(group_name: str, limit: int = 10) -> List[Dict[str, Any]]:
    session = SessionLocal()
    try:
        rows = (session.query(
                    MarketBrandORM.brand,
                    MarketBrandORM.company,
                    MarketBrandORM.molecule_desc,
                    func.sum(MarketBrandORM.value_latest).label("value_latest"),
                    func.sum(MarketBrandORM.value_prev).label("value_prev"),
                )
                .filter(MarketBrandORM.group_name == group_name)
                .group_by(MarketBrandORM.brand, MarketBrandORM.company)
                .order_by(func.sum(MarketBrandORM.value_latest).desc())
                .limit(limit).all())
        return [{
            "brand": r.brand,
            "company": r.company,
            "molecule_desc": r.molecule_desc,
            "value_latest": round(r.value_latest or 0.0, 2),
            "growth_percent": _growth(r.value_latest, r.value_prev),
        } for r in rows]
    finally:
        session.close()


def search_brands(query: str, limit: int = 30) -> List[Dict[str, Any]]:
    """Free-text lookup across brand, molecule, and company."""
    text = (query or "").strip()
    if len(text) < 2:
        return []
    pattern = f"%{text.upper()}%"
    session = SessionLocal()
    try:
        rows = (session.query(
                    MarketBrandORM.brand,
                    MarketBrandORM.company,
                    MarketBrandORM.molecule_desc,
                    MarketBrandORM.group_name,
                    func.sum(MarketBrandORM.value_latest).label("value_latest"),
                    func.sum(MarketBrandORM.value_prev).label("value_prev"),
                )
                .filter(or_(
                    func.upper(MarketBrandORM.brand).like(pattern),
                    func.upper(MarketBrandORM.molecule_desc).like(pattern),
                    func.upper(MarketBrandORM.company).like(pattern),
                ))
                .group_by(MarketBrandORM.brand, MarketBrandORM.company)
                .order_by(func.sum(MarketBrandORM.value_latest).desc())
                .limit(limit).all())
        return [{
            "brand": r.brand,
            "company": r.company,
            "molecule_desc": r.molecule_desc,
            "group_name": r.group_name,
            "value_latest": round(r.value_latest or 0.0, 2),
            "growth_percent": _growth(r.value_latest, r.value_prev),
        } for r in rows]
    finally:
        session.close()


def molecule_overview(molecule: str) -> Dict[str, Any]:
    """One call for the market panel: size, brands, companies, class rivals.

    `companies` is the display leaderboard and is capped; `total_companies` is
    the real count. Callers must not infer one from the other — reading
    len(companies) as the total reported 15 companies for a molecule with 149.
    """
    competitors = brand_competitors(molecule)
    leaderboard = company_leaderboard(molecule)
    return {
        **competitors,
        "companies": leaderboard,
        "total_companies": count_companies(molecule),
        "class": class_competitors(molecule),
        "datasets": list_datasets(),
        "has_data": bool(competitors.get("brands")),
    }


def count_companies(molecule: str) -> int:
    """How many companies market this molecule, ignoring any display cap."""
    condition, _ = _molecule_filter(molecule)
    if condition is None:
        return 0
    session = SessionLocal()
    try:
        dataset_id = active_dataset_id(session, condition)
        if dataset_id is None:
            return 0
        return (session.query(func.count(func.distinct(MarketBrandORM.company)))
                .filter(condition,
                        MarketBrandORM.dataset_id == dataset_id,
                        MarketBrandORM.company.isnot(None))
                .scalar() or 0)
    finally:
        session.close()


# --------------------------------------------------------------------------
# Manual competitors
#
# A licensed extract answers "who competed as of the file's period" — it can
# be silent on a brand that launched or scaled after that date, or on any
# market no extract has ever been loaded for. This lets a team record a
# competitor it knows is real, with its own source, without ever presenting
# that attestation as if it carried the same audit weight as a licensed row.
# --------------------------------------------------------------------------

def add_manual_competitor(
    molecule: str,
    brand: str,
    source_note: str,
    added_by: str,
    company: Optional[str] = None,
    market: Optional[str] = None,
    value_estimate: Optional[float] = None,
    value_unit: Optional[str] = None,
    value_basis: Optional[str] = None,
    mrp: Optional[float] = None,
    ptr: Optional[float] = None,
    pts: Optional[float] = None,
    price_unit: Optional[str] = None,
) -> Dict[str, Any]:
    """Record a team-attested competitor. `source_note` is mandatory — an
    entry with no stated source is exactly the unsourced claim this
    application exists to prevent.

    mrp/ptr/pts here describe THIS competitor brand's own trade price, not
    the caller's — distinct from the forecast module's trade price structure,
    which models the user's own planned brand. MRP is the only one of the
    three a public source can ever carry (a retail listing); PTR and PTS are
    confidential terms in the competitor's own distribution agreements, so
    they will almost always be blank unless a team's own trade contacts
    supplied one — never populated from a scrape or an estimate.
    """
    if not source_note or not source_note.strip():
        raise ValueError("A manual competitor entry must state its source.")
    if (ptr is not None or pts is not None) and mrp is None:
        raise ValueError(
            "PTR/PTS without an MRP is unusual enough to be worth double-checking — "
            "supply the MRP too, or omit PTR/PTS if it genuinely is not known."
        )

    record = ManualCompetitorORM(
        id=uuid.uuid4().hex,
        molecule_key=molecule_search_key(molecule),
        molecule_desc=molecule.strip(),
        brand=brand.strip(),
        company=(company or "").strip() or None,
        market=(market or "").strip() or None,
        value_estimate=value_estimate,
        value_unit=(value_unit or "").strip() or None,
        value_basis=(value_basis or "").strip() or None,
        mrp=mrp,
        ptr=ptr,
        pts=pts,
        price_unit=(price_unit or "").strip() or None,
        source_note=source_note.strip(),
        added_by=added_by.strip() or "Unattributed",
        added_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )
    session = SessionLocal()
    try:
        session.add(record)
        session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()
    return _manual_competitor_dict(record)


def _manual_competitor_dict(row) -> Dict[str, Any]:
    return {
        "id": row.id,
        "molecule_desc": row.molecule_desc,
        "brand": row.brand,
        "company": row.company,
        "market": row.market,
        "value_estimate": row.value_estimate,
        "value_unit": row.value_unit,
        "value_basis": row.value_basis,
        "mrp": row.mrp,
        "ptr": row.ptr,
        "pts": row.pts,
        "price_unit": row.price_unit,
        "source_note": row.source_note,
        "added_by": row.added_by,
        "added_at": row.added_at,
    }


def list_manual_competitors(molecule: str) -> List[Dict[str, Any]]:
    key = molecule_search_key(molecule)
    if not key:
        return []
    session = SessionLocal()
    try:
        rows = (session.query(ManualCompetitorORM)
                .filter(ManualCompetitorORM.molecule_key.like(f"%{key}%"))
                .order_by(ManualCompetitorORM.added_at.desc()).all())
        return [_manual_competitor_dict(r) for r in rows]
    finally:
        session.close()


def delete_manual_competitor(entry_id: str) -> bool:
    session = SessionLocal()
    try:
        row = session.get(ManualCompetitorORM, entry_id)
        if not row:
            return False
        session.delete(row)
        session.commit()
        return True
    finally:
        session.close()
